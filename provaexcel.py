from openpyxl import Workbook, load_workbook

wb = load_workbook('sample.xlsx')
sheet = wb.active
dataAtA1 = sheet['A2']
dataAtA2 = sheet.cell(row=3, column=3)
print('Data at A2:', dataAtA1.value)
print('Data at C3:', dataAtA2.value)
dataProva = sheet['A1']



wb = Workbook()

# grab the active worksheet
ws = wb.active

<<<<<<< HEAD
# Data can be assigned directly to cells
ws['A1'] = 44
=======
>>>>>>> e038ef4b6d76f2f8020233035b1965e6af4b6e53


# Data can be assigned directly to cells
ws['A1'] = 'Edoardo'
ws['B1'] = 'Enrico'
ws['C1'] = 'Marco'
ws['D1'] = 'Alessandro'


print(ws['A1'])

if(dataProva.value != None):
# Rows can also be appended
<<<<<<< HEAD
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([1, 2, 3])
ws.append([1, 2, 3])
=======
    ws.append([2, 2, 4, 5, 6, 7, 8, 9, 10])
    ws.append([1, 2, 4, 5, 6, 7, 8, 9, 10])
    ws.append([1, 2, 4, 5, 6, 7, 8, 9, 10])
    ws.append([1, 2, 4, 5, 6, 7, 8, 9, 10])
    ws.append([2, 2, 4, 5, 6, 7, 8, 9, 10])
    ws.append([1, 2, 4, 5, 6, 7, 8, 9, 10])

ws['K1'] = 'Edoardo'
ws['B10'] = 'Enrico'
ws['E11'] = 'Marco'
ws['G13'] = 'Alessandro'
>>>>>>> e038ef4b6d76f2f8020233035b1965e6af4b6e53

# Python types will automatically be converted
import datetime
ws['A3'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")